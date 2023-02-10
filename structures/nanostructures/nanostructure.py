import logging
from copy import copy
from dataclasses import field, dataclass
from typing import List

import xlsxwriter
from openpyxl import load_workbook

from structures.domains import Domains
from structures.helices import DoubleHelices
from structures.profiles import NucleicAcidProfile
from structures.strands import Strands

logger = logging.getLogger(__name__)


@dataclass
class Nanostructure:
    """
    An object representing an entire nanostructure.

    Attributes:
        strands: The strands in the nanostructure.
        domains: The domains in the nanostructure.
        nucleic_acid_profile: The nucleic acid settings that the nanostructure uses.
    """

    strands: field(default_factory=Strands)
    domains: field(default_factory=Domains)
    nucleic_acid_profile: field(default_factory=NucleicAcidProfile)

    def to_file(
        self,
        filepath: str,
        additional_nucleic_acid_profiles: List[NucleicAcidProfile] | tuple = (),
    ) -> None:
        """
        Create a multipage spreadsheet with the current nanostructure.

        Args:
            filepath: The path to the file to save.
            additional_nucleic_acid_profiles: Additional nucleic acid profiles to
                include in the save. If this is an empty tuple only the current
                nucleic acid profile is used.
        """
        logger.debug("Saving nanostructure to file: %s", filepath)

        workbook = xlsxwriter.Workbook(filepath)

        current_nucleic_acid_profile = copy(self.nucleic_acid_profile)
        current_nucleic_acid_profile.name = "Computed Profile"
        nucleic_acid_profiles = [current_nucleic_acid_profile]
        nucleic_acid_profiles.extend(additional_nucleic_acid_profiles)

        self.nucleic_acid_profile.write_worksheet(
            workbook, profiles=nucleic_acid_profiles
        )
        self.domains.write_worksheet(workbook)
        self.strands.write_worksheets(workbook)

        workbook.close()

        logger.debug("Saved nanostructure")

    @classmethod
    def from_file(cls, filepath: str) -> "Nanostructure":
        workbook = load_workbook(filepath)

        return cls(
            Strands.read_worksheet(workbook["Strands"]),
            Domains.read_worksheet(workbook["Domains"]),
            NucleicAcidProfile.read_worksheet(workbook["Nucleic Acid Profile"]),
        )
